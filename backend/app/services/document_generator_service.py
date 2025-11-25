"""
Сервис для генерации документов из шаблонов
Поддерживает генерацию в форматах PDF и Excel (XLSX)
"""
import logging
import os
import tempfile
import asyncio
from pathlib import Path
from typing import Dict, Any, Optional
from datetime import datetime
from jinja2 import Template, Environment, FileSystemLoader
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class DocumentGeneratorService:
    """Сервис для генерации документов из шаблонов"""
    
    def __init__(self):
        self.base_dir = Path(__file__).parent.parent.parent
        self.templates_dir = self.base_dir / "templates" / "documents"
        self.generated_dir = self.base_dir / "uploads" / "generated_documents"
        
        # Создаем директории если их нет
        try:
            self.templates_dir.mkdir(parents=True, exist_ok=True)
        except PermissionError:
            logger.warning(f"Не удалось создать директорию шаблонов: {self.templates_dir}")
        
        try:
            self.generated_dir.mkdir(parents=True, exist_ok=True)
        except PermissionError:
            logger.warning(f"Не удалось создать директорию для генерации: {self.generated_dir}")
            # Используем временную директорию как fallback
            import tempfile
            self.generated_dir = Path(tempfile.gettempdir()) / "generated_documents"
            self.generated_dir.mkdir(parents=True, exist_ok=True)
        
        # Jinja2 окружение для HTML шаблонов
        self.jinja_env = Environment(
            loader=FileSystemLoader(str(self.templates_dir)),
            autoescape=True
        )
    
    async def generate_document(
        self,
        template_id: int,
        template_file: str,
        template_type: str,
        data: Dict[str, Any],
        output_format: str = "pdf",
        user_id: int = None
    ) -> Dict[str, Any]:
        """
        Генерирует документ из шаблона
        
        Args:
            template_id: ID шаблона
            template_file: Путь к файлу шаблона
            template_type: Тип шаблона ("html" или "xlsx")
            data: Данные для заполнения
            output_format: Формат выходного файла ("pdf" или "xlsx")
            user_id: ID пользователя
            
        Returns:
            Словарь с путем к файлу и метаданными
        """
        try:
            # Добавляем системные данные
            data_with_system = {
                **data,
                "current_date": datetime.now().strftime("%d.%m.%Y"),
                "current_time": datetime.now().strftime("%H:%M"),
                "current_datetime": datetime.now().strftime("%d.%m.%Y %H:%M")
            }
            
            if template_type == "html" and output_format == "pdf":
                # Генерация PDF из HTML шаблона
                file_path = await self._generate_pdf_from_html(template_file, data_with_system, user_id)
            elif template_type == "xlsx" and output_format == "xlsx":
                # Генерация Excel из XLSX шаблона
                file_path = await self._generate_xlsx_from_template(template_file, data_with_system, user_id)
            else:
                raise ValueError(f"Неподдерживаемая комбинация: template_type={template_type}, output_format={output_format}")
            
            # Получаем размер файла
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            
            return {
                "file_path": str(file_path),
                "file_size": file_size,
                "file_format": output_format,
                "filename": os.path.basename(file_path)
            }
            
        except Exception as e:
            logger.error(f"Ошибка генерации документа: {e}")
            raise
    
    async def _generate_pdf_from_html(
        self,
        template_file: str,
        data: Dict[str, Any],
        user_id: Optional[int] = None
    ) -> str:
        """
        Генерирует PDF из HTML шаблона
        
        Args:
            template_file: Путь к HTML шаблону
            data: Данные для заполнения
            user_id: ID пользователя
            
        Returns:
            Путь к сгенерированному PDF файлу
        """
        try:
            # Загружаем HTML шаблон
            template_path = self.templates_dir / template_file
            if not template_path.exists():
                raise FileNotFoundError(f"Шаблон не найден: {template_path}")
            
            with open(template_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            # Заполняем шаблон данными через Jinja2
            template = Template(html_content)
            filled_html = template.render(**data)
            
            # Конвертируем HTML в PDF
            try:
                from weasyprint import HTML
                
                # Генерируем имя файла
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"document_{timestamp}_{user_id or 'anonymous'}.pdf"
                output_path = self.generated_dir / filename
                
                # Генерируем PDF
                await asyncio.to_thread(
                    HTML(string=filled_html).write_pdf,
                    str(output_path)
                )
                
                return str(output_path)
                
            except ImportError:
                # Fallback на reportlab если weasyprint не установлен
                logger.warning("weasyprint не установлен, используем reportlab")
                return await self._generate_pdf_with_reportlab(filled_html, user_id)
                
        except Exception as e:
            logger.error(f"Ошибка генерации PDF: {e}")
            raise
    
    async def _generate_pdf_with_reportlab(
        self,
        html_content: str,
        user_id: Optional[int] = None
    ) -> str:
        """
        Генерирует PDF используя reportlab (fallback)
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from html import unescape
            import re
            
            # Простой парсинг HTML в reportlab элементы
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"document_{timestamp}_{user_id or 'anonymous'}.pdf"
            output_path = self.generated_dir / filename
            
            doc = SimpleDocTemplate(str(output_path), pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Убираем HTML теги и создаем параграфы
            text = re.sub(r'<[^>]+>', '', html_content)
            text = unescape(text)
            
            for line in text.split('\n'):
                if line.strip():
                    story.append(Paragraph(line.strip(), styles['Normal']))
                    story.append(Spacer(1, 12))
            
            await asyncio.to_thread(doc.build, story)
            
            return str(output_path)
            
        except ImportError:
            raise ImportError("Необходимо установить weasyprint или reportlab для генерации PDF")
    
    async def _generate_xlsx_from_template(
        self,
        template_file: str,
        data: Dict[str, Any],
        user_id: Optional[int] = None
    ) -> str:
        """
        Генерирует Excel файл из XLSX шаблона
        
        Args:
            template_file: Путь к XLSX шаблону
            data: Данные для заполнения
            user_id: ID пользователя
            
        Returns:
            Путь к сгенерированному XLSX файлу
        """
        try:
            # Загружаем XLSX шаблон
            template_path = self.templates_dir / template_file
            if not template_path.exists():
                raise FileNotFoundError(f"Шаблон не найден: {template_path}")
            
            # Загружаем workbook
            workbook = await asyncio.to_thread(load_workbook, str(template_path))
            
            # Заполняем плейсхолдеры в каждой ячейке
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            # Заменяем плейсхолдеры {{variable}} на значения
                            value = cell.value
                            for key, val in data.items():
                                placeholder = f"{{{{{key}}}}}"
                                if placeholder in value:
                                    value = value.replace(placeholder, str(val))
                            cell.value = value
            
            # Сохраняем файл
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"document_{timestamp}_{user_id or 'anonymous'}.xlsx"
            output_path = self.generated_dir / filename
            
            await asyncio.to_thread(workbook.save, str(output_path))
            
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Ошибка генерации Excel: {e}")
            raise
    
    def validate_template_data(
        self,
        required_fields: list,
        input_data: Dict[str, Any]
    ) -> tuple[bool, list[str]]:
        """
        Валидирует данные для шаблона
        
        Args:
            required_fields: Список обязательных полей
            input_data: Входные данные
            
        Returns:
            (is_valid, list_of_errors)
        """
        if not required_fields:
            return True, []
        
        errors = []
        for field in required_fields:
            field_name = field.get('name') or field.get('key')
            if not field_name:
                continue
            
            if field_name not in input_data or not input_data[field_name]:
                field_label = field.get('label', field_name)
                errors.append(f"Поле '{field_label}' обязательно для заполнения")
        
        return len(errors) == 0, errors


# Singleton instance
document_generator_service = DocumentGeneratorService()

